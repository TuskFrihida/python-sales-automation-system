"""
Excel report automation.

Turns the sales data into a polished, multi-sheet ``.xlsx`` workbook with
styled headers, KPI summary, raw data, aggregated tables and *native* Excel
charts (bar, pie and line). Native charts stay interactive inside Excel,
unlike the static PNGs produced by matplotlib.

This is the kind of deliverable clients love: a single file they can open,
filter and present without touching Python.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from report_generator import SalesReportGenerator
from utils import ensure_folder

# Reusable style constants -----------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="2E86AB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="2E86AB")
LABEL_FONT = Font(bold=True, color="444444")
CURRENCY_FMT = '"$"#,##0.00'
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


class ExcelReportExporter:
    """Builds a formatted Excel workbook from the sales data."""

    def __init__(self, data_folder: str = "data", reports_folder: str = "reports"):
        self.reports_folder = ensure_folder(reports_folder)
        self.generator = SalesReportGenerator(data_folder, reports_folder)

    # -- public API ------------------------------------------------------------
    def export(self, filename: str | None = None) -> Path:
        """Generate the workbook and return its path."""
        daily_df, source_name = self.generator.load_latest_data()
        weekly_df = self.generator.load_weekly_data()
        stats = self.generator.generate_summary_stats(daily_df)

        wb = Workbook()
        self._build_summary_sheet(wb.active, stats, source_name)
        self._build_aggregate_sheet(wb.create_sheet("By Product"), daily_df,
                                    group="Product", chart="bar")
        self._build_aggregate_sheet(wb.create_sheet("By Region"), daily_df,
                                    group="Region", chart="pie")
        self._build_trend_sheet(wb.create_sheet("Daily Trend"), weekly_df)
        self._build_raw_sheet(wb.create_sheet("Raw Data"), daily_df)

        if filename is None:
            filename = f"sales_report_{datetime.now():%Y-%m-%d}.xlsx"
        path = self.reports_folder / filename
        wb.save(path)
        print(f"Excel report saved: {path}")
        return path

    # -- sheet builders --------------------------------------------------------
    def _build_summary_sheet(self, ws, stats: dict, source_name: str) -> None:
        ws.title = "Summary"
        ws["A1"] = "Sales Performance Summary"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Generated {datetime.now():%B %d, %Y at %I:%M %p}"
        ws["A3"] = f"Source: {source_name}"

        rows = [
            ("Total Revenue", stats["total_revenue"], CURRENCY_FMT),
            ("Total Transactions", stats["total_transactions"], None),
            ("Average Order Value", stats["average_order_value"], CURRENCY_FMT),
            ("Top Product", stats["top_product"], None),
            ("Top Salesperson", stats["top_salesperson"], None),
            ("Top Region", stats["top_region"], None),
            ("Date Range", stats["date_range"], None),
        ]
        start = 5
        for offset, (label, value, fmt) in enumerate(rows):
            r = start + offset
            ws.cell(r, 1, label).font = LABEL_FONT
            cell = ws.cell(r, 2, value)
            if fmt:
                cell.number_format = fmt
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 28

    def _build_aggregate_sheet(self, ws, df: pd.DataFrame, group: str, chart: str) -> None:
        agg = (df.groupby(group)["Total_Amount"].sum()
               .sort_values(ascending=False).reset_index())
        agg.columns = [group, "Revenue"]
        self._write_table(ws, agg)

        if chart == "bar":
            obj = BarChart()
            obj.title = f"Revenue by {group}"
            obj.y_axis.title = "Revenue ($)"
            obj.x_axis.title = group
        else:
            obj = PieChart()
            obj.title = f"Revenue distribution by {group}"

        data = Reference(ws, min_col=2, min_row=1, max_row=len(agg) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(agg) + 1)
        obj.add_data(data, titles_from_data=True)
        obj.set_categories(cats)
        obj.height = 9
        obj.width = 18
        ws.add_chart(obj, "E2")

    def _build_trend_sheet(self, ws, df: pd.DataFrame) -> None:
        trend = df.groupby(df["Date"].dt.date)["Total_Amount"].sum().reset_index()
        trend.columns = ["Date", "Revenue"]
        trend["Date"] = trend["Date"].astype(str)
        self._write_table(ws, trend)

        chart = LineChart()
        chart.title = "Daily Revenue Trend"
        chart.y_axis.title = "Revenue ($)"
        chart.x_axis.title = "Date"
        data = Reference(ws, min_col=2, min_row=1, max_row=len(trend) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(trend) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 9
        chart.width = 18
        ws.add_chart(chart, "E2")

    def _build_raw_sheet(self, ws, df: pd.DataFrame) -> None:
        export_df = df.copy()
        export_df["Date"] = export_df["Date"].astype(str)
        self._write_table(ws, export_df)
        # Format currency columns when present.
        for col_idx, name in enumerate(export_df.columns, start=1):
            if name in ("Unit_Price", "Total_Amount"):
                letter = get_column_letter(col_idx)
                for cell in ws[letter][1:]:
                    cell.number_format = CURRENCY_FMT

    # -- helpers ---------------------------------------------------------------
    @staticmethod
    def _write_table(ws, df: pd.DataFrame) -> None:
        """Write a DataFrame with a styled header row and auto-sized columns."""
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(r_idx, c_idx, value)
                cell.border = BORDER
                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = CENTER
        for c_idx, name in enumerate(df.columns, start=1):
            longest = max([len(str(name))] + [len(str(v)) for v in df.iloc[:, c_idx - 1]])
            ws.column_dimensions[get_column_letter(c_idx)].width = min(longest + 4, 40)


def main() -> None:
    ExcelReportExporter().export()


if __name__ == "__main__":
    main()
